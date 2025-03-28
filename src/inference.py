import joblib
import pandas as pd
from openpyxl import load_workbook

print("Loading in models")
post_fight_model = joblib.load('models\\post_fight_model.joblib')
win_pred_model = joblib.load('models\\win_pred_model.joblib')
print("Models loaded successfully.")

# Column names
pre_features_numeric = [
    "r_wins_agg", "r_losses_agg", "r_height_agg", "r_weight_agg", "r_reach_agg", 
    "r_age_agg", "r_SLpM_agg", "r_sig_str_acc_agg", "r_SApM_agg", "r_str_def_agg",
    "r_td_avg_agg", "r_td_acc_agg", "r_td_def_agg", "r_sub_avg_agg", 
    "b_wins_agg", "b_losses_agg", "b_height_agg", "b_weight_agg", "b_reach_agg",
    "b_age_agg", "b_SLpM_agg", "b_sig_str_acc_agg", "b_SApM_agg", "b_str_def_agg",
    "b_td_avg_agg", "b_td_acc_agg", "b_td_def_agg", "b_sub_avg_agg"
]
pre_features_categorical = ["weight_class", "is_title_bout", "gender", "r_stance_agg", "b_stance_agg"]

post_features_numeric = [
    "finish_round", "total_rounds", "time_sec", 
    "r_kd", "r_sig_str", "r_sig_str_att", "r_sig_str_acc", "r_str",
    "r_str_att", "r_str_acc", "r_td", "r_td_att", "r_td_acc",
    "r_sub_att", "r_rev", "r_ctrl_sec", 
    "b_kd", "b_sig_str", "b_sig_str_att", "b_sig_str_acc", "b_str",
    "b_str_att", "b_str_acc", "b_td", "b_td_att", "b_td_acc",
    "b_sub_att", "b_rev", "b_ctrl_sec"
]

col_ordering = [
    "wins", "losses", "height", "weight", "reach", "age", 
    "SLpM", "sig_str_acc", "SApM", "str_def", 
    "td_avg", "td_acc", "td_def", "sub_avg"
]

weight_classes = {
    "Fl" : "Flyweight",
    "Bw" : "Bantamweight",
    "Ft" : "Featherweight",
    "Lw" : "Lightweight",
    "Ww" : "Welterweight",
    "Mw" : "Middleweight",
    "Lh" : "Light Heavyweight",
    "Hw" : "Heavyweight",
    "Ws" : "Women's Strawweight",
    "Wf" : "Women's Flyweight",
    "Wb" : "Women's Bantamweight",
    "Cw" : "Catch Weight"
}

# Read in data for automation
fighters = pd.read_csv("data\\fighter_stats.csv")

# Read in spreadsheet to log predictions
predictions = pd.read_excel("Predictions.xlsx")
wb = load_workbook("Predictions.xlsx")
ws = wb["New"]

for i, row in enumerate(predictions.itertuples(index=False), start = 2):

    fighter_1 = row.Fighter_1
    fighter_2 = row.Fighter_2
    weight = row.Weight
    is_title = [int(row.Title)]
    gender = [row.Gender]

    fighter_1_stats = fighters[fighters["name"] == fighter_1]
    fighter_2_stats = fighters[fighters["name"] == fighter_2]

    if fighter_1_stats.empty or fighter_2_stats.empty:

        if fighter_1_stats.empty:
            print(f"⚠️ Warning: Fighter '{fighter_1}' not found in dataset. Skipping row {i}.")

        if fighter_2_stats.empty:
            print(f"⚠️ Warning: Fighter '{fighter_2}' not found in dataset. Skipping row {i}.")

        continue  # Skip to the next iteration

    weight = [weight_classes[weight]]

    # List formatted
    fighters_concat = fighter_1_stats[col_ordering].iloc[0].tolist() + fighter_2_stats[col_ordering].iloc[0].tolist() + weight + is_title + gender + [fighter_1_stats["stance"].iloc[0]] + [fighter_2_stats["stance"].iloc[0]]

    # DataFrame formatted for input into model
    fight = pd.DataFrame([fighters_concat], columns = pre_features_numeric + pre_features_categorical)

    print("Calculating post fight stats predictions")
    post_fight_predictions = post_fight_model.predict(fight)
    print(post_fight_predictions)

    post_fight_predictions = pd.DataFrame(post_fight_predictions, columns=post_features_numeric)
    combined_input = pd.concat([fight, post_fight_predictions], axis=1)

    print("Calculating win probabilities")
    win_probabilities = win_pred_model.predict_proba(combined_input)
    print(f"{fighter_1} Win Probability: {win_probabilities[0][0]}\n{fighter_2} Win Probability: {win_probabilities[0][1]}\n")

    ws.cell(row=i, column=6, value=f'{win_probabilities[0][0]}')
    ws.cell(row=i, column=7, value=f'{win_probabilities[0][1]}')
    
wb.save("Predictions.xlsx")
